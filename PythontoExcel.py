import openpyxl as xl
from openpyxl.styles import Font

#Create a new excel document
wb = xl.Workbook()

#Create sheet to write to
ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1,title='Second Sheet')

#Write content to a cell
ws["A1"] = 'Invoice'
headerFont = Font(name='Times New Roman',size=24,bold=True)
ws["A1"].font = headerFont

ws["A2"] = 'Tires'
ws["A3"] = 'Brakes'
ws["A4"] = 'Alignment'

ws.merge_cells('A1:B1')

ws["B2"] = 450
ws["B3"] = 225
ws["B4"] = 150

#To unmerge cells
#ws.unmerge_cells('A1:B1')

ws["A8"] = 'Total'
ws["A8"].font = Font(size=16,bold=True)

ws["B8"] = '=SUM(B2:B4)'

ws.column_dimensions["A"].wdith = 25



font = Font(name='Times New Roman',size=16,bold=True)

write_sheet = wb['Second Sheet']
read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

# My code
# for row in range(1,read_ws.max_row+1):
#     for col in range(1, 5): 
#         write_sheet.cell(row=row, column=col).value = read_ws.cell(row=row, column=col).value

# write_sheet["B43"] = 'Total'
# write_sheet["B43"].font = font

# write_sheet["C43"] = '=SUM(C2:C41)'

# write_sheet["D43"] = '=SUM(D2:D41)'

# write_sheet["B45"] = 'Avgerage'
# write_sheet["B45"].font = font

# write_sheet["C45"] = '=AVERAGE(C2:C41)'

# write_sheet["D45"] = '=AVERAGE(D2:D41)'


# Professor Code
for row in read_ws.iter_rows():
    ls = [i.value for i in row]
    write_sheet.append(ls)

max_row = write_sheet.max_row

write_sheet.cell(max_row+2,2).value = 'Total'
write_sheet.cell(max_row+2,2).font = font

write_sheet.cell(max_row+2,3).value = '=SUM(C2:C' + str(max_row) + ')'
write_sheet.cell(max_row+2,4).value = '=SUM(D2:D' + str(max_row) + ')'

write_sheet.cell(max_row+4,2).value = 'Average'
write_sheet.cell(max_row+4,2).font = font

write_sheet.cell(max_row+4,3).value = '=AVERAGE(C2:C' + str(max_row) + ')'
write_sheet.cell(max_row+4,4).value = '=AVERAGE(D2:D' + str(max_row) + ')'

write_sheet.column_dimensions["A"].width = 16
write_sheet.column_dimensions["B"].width = 15
write_sheet.column_dimensions["C"].width = 15
write_sheet.column_dimensions["D"].width = 15

for cell in write_sheet["C:C"]:
    cell.number_format = '#,##0'

for cell in write_sheet["D:D"]:
    cell.number_format = u'"$ "#,##0.00'


wb.save("PythontoExcel.xlsx")