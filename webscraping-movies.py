
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			
soup = BeautifulSoup(page, 'html.parser')
title = soup.title
print(title.text)

table_rows=soup.findAll("tr")

movies = []
for row in table_rows[1:6]:
    td=row.findAll("td")
    rank=td[0].text
    movie=td[1].text    
    release_date=td[8].text
    num_theaters=float(td[6].text.replace(",", ""))
    gross=float(td[5].text.strip("$").replace(",", ""))
    
    avg_gross = gross/num_theaters

    movie_info = {
        "Rank": rank,
        "Movie": movie,
        "Release Date": release_date,
        "Num of Theaters": num_theaters,
        "Total Gross": gross,
        "Avg Gross": avg_gross
    }
    movies.append(movie_info)

wb = xl.Workbook()
ws = wb.active
ws.title = 'Box Office Report'

#Header
headerFont = Font(name='Times New Roman',size=24,bold=True)
ws["A1"] = 'No.'
ws["A1"].font = headerFont
ws["B1"] = 'Movie Title'
ws["B1"].font = headerFont
ws["C1"] = 'Release Date'
ws["C1"].font = headerFont
ws["D1"] = 'Number of Theaters'
ws["D1"].font = headerFont
ws["E1"] = 'Total Gross'
ws["E1"].font = headerFont
ws["F1"] = 'Average Gross by Theater'
ws["F1"].font = headerFont

# for row in range(1,ws.max_row+1):
#     for cell in ws["A:F"]:
#         ws.cell(movies)

# We didn't finish this one in class








